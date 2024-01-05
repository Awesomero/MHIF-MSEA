import os
import openpyxl
import pandas as pd
from pyrwr.pyrwr.ppr import PPR

# Obtain the user-submitted list of differentially expressed miRNAs.
mirna_list = pd.read_excel('data/miRNA_lists/colon_list.xlsx', header=None, names=['miRNA'], engine='openpyxl')
miRNA_list = mirna_list['miRNA'].tolist()

# user_list_num represents the number of lncRNAs in the differentially expressed miRNA list submitted by the user.
user_list_num = len(mirna_list)

# Step 1----------Generate the intersection of differentially expressed lncRNAs and all miRNAs-----------
wb1 = openpyxl.load_workbook('data/mirna_names/miRNA_name_average_dis_go.xlsx')  # 取出全部miRNA
sh1 = wb1['Sheet1']
# Initialize the list for the intersection of lncRNAs.
intersection_list = []
# Find the intersection between the user-submitted list and all lncRNAs.
for i in range(user_list_num):
    for j in range(1041):
        if mirna_list['miRNA'][i] == sh1.cell(j + 1, 1).value:
            intersection_list.append(mirna_list['miRNA'][i])
print(intersection_list)
intersection_list_num = len(intersection_list)

# Step 2----------------------Generate indices for the intersection of lncRNAs.--------------------
xuhao_path = 'data/miRNA_index/Colon_Carcinoma/mirna_xuhao_colon_average_dis_go.txt'
# If the index file already exists, delete it and create a new one.
if os.path.exists(xuhao_path):
    os.remove(xuhao_path)

for i in range(intersection_list_num):
    for j in range(1041):
        if sh1.cell(j + 1, 1).value == intersection_list[i]:
            with open('data/miRNA_index/Colon_Carcinoma/mirna_xuhao_colon_average_dis_go.txt', 'a') as f:
                f.write(str(sh1.cell(j + 1, 2).value) + ' ')


# Step 3----------------------Random Walk Expansion of lncRNA List--------------------
def produce_d(seeds):
    seeds1 = seeds  # Seed sequence
    c = 0.15  # Restart probability
    epsilon = 1e-9  # Value for stopping iteration when the error is less than epsilon
    max_iters = 100  # Maximum number of iterations
    n = 10  # n is the number of miRNAs to output
    ppr = PPR()
    ppr.read_graph('./get_edge_lists/edge_average_dis_go/0.60.txt', graph_type='directed')  # Directed Graph
    r = ppr.compute(seeds1, c, epsilon, max_iters)
    dict = {}

    for i in range(len(r)):  # Create a dictionary to retain indices
        dict[i + 1] = r[i]

    d_order = sorted(dict.items(), key=lambda x: x[1], reverse=True)  # Sort by probability
    return d_order


def RW_list(seeds, n):
    result = []

    d = produce_d(seeds)

    for i in range(min(n, len(d))):
        if d[i][1] > 0:  # Exclude probabilities with zero in random walk
            result.append(d[i])
        else:
            result.append((i, 0))
    return result


with open("data/miRNA_index/Colon_Carcinoma/mirna_xuhao_colon_average_dis_go.txt", "r") as f:
    data = f.readline()
    data1 = data.split(' ')
    data1.pop()
    data1 = [int(i) for i in data1]

    num = 1041
    a = RW_list(data1, num)
    print(a)
    print(len(a))

    extended_list = []  # Initialize the list after random walk expansion

    for i in range(797):
        if a[i][1] == 0:
            continue
        for j in range(1041):
            if sh1.cell(j + 1, 2).value == a[i][0]:
                extended_list.append(sh1.cell(j + 1, 1).value)  # List of miRNAs containing the intersection lncRNAs
                # print(extended_list)
    # Initialize the list of additionally expanded miRNAs
    extended_list1 = []
    for mi1 in extended_list:  # Filter out the additionally expanded list
        if mi1 in extended_list and mi1 not in intersection_list:
            extended_list1.append(mi1)
    # Add the additionally expanded list to the user-input miRNA list
    for i in extended_list1:
        miRNA_list.append(i)

    # Convert data to DataFrame
    df = pd.DataFrame({"miRNA": miRNA_list})

    # Write DataFrame to Excel file
    output_file = "data/extended_lists/Colon_Carcinoma/expend_miRNA_data_colon_average_dis_go.xlsx"
    df.to_excel(output_file, index=False)
